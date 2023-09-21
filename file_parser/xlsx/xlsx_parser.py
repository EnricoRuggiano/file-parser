from ..abstract import ParserInterface
from ..exceptions import *
from ..commons import *
from typing import Mapping

from io import BytesIO
import openpyxl

class XlsxParser(ParserInterface):      
    format = 'xlsx'
    worksheet_index:int = 0
    body:List[List[str]] = []

    def parse_input(self) -> List[Mapping]:
        check_is_not_null(self.raw_content)
        check_is_bytes(self.raw_content)

        f:BytesIO = BytesIO(self.raw_content)
        xlsx_file = openpyxl.load_workbook(f)

        # get the worksheet
        ws = xlsx_file.worksheets[self.worksheet_index]

        # reset
        self.body = []
        self.headers = []
        
        # get headers
        line:int = 1
        for y in range(1, ws.max_column + 1):
            self.headers.append(str(ws.cell(1, y).value).strip())
        line += 1

        if len(self.headers) == 0:
            raise EmptyHeadersException()

        # get data
        for x in range(2, ws.max_row + 1):
            r:List = []
            for y in range(1, ws.max_column + 1):
                r.append(str(ws.cell(x, y).value).strip())
            self.body.append(r)
            line += 1

        print(self.body)
        return self.__xlsx_to_dict(self.headers, self.body)
    
    def parse_output(self):
        pass

    @classmethod
    def __xlsx_to_dict(self, headers:List, content:List)->List[Mapping]:
        out:List[Mapping] = []
        line:int = 2
        for x in content:
            d:Mapping = dict()
            if not x:
                pass
            else:
                try:
                    for y in range(0, len(headers)):
                        d[headers[y]] = x[y]
                    out.append(d)
                except Exception as e:
                    print("Impossible to parse line {}".format(line))
                    print(e)
            line += 1
        return out
